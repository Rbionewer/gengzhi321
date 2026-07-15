#!/usr/bin/env python3
"""Prepare NetAct input files from transcriptome annotation workbooks."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


TF_KEYWORDS = (
    "transcription factor",
    "transcriptional activator",
    "transcriptional repressor",
    "myb",
    "bhlh",
    "wrky",
    "nac",
    "bzip",
    "mads",
    "ap2/erf",
    "erf",
    "arf",
    "tcp",
    "gata",
    "gras",
    "wox",
    "yabby",
    "nf-y",
    "hd-zip",
    "trihelix",
    "sbp",
    "spl",
    "hsf",
)

TF_SYMBOL_RE = re.compile(
    r"^(myb|bhlh|wrky|nac|bzip|mads|erf|arf|tcp|gata|spl|hsf|nfy|wox|trihelix)",
    re.IGNORECASE,
)
KO_ENTRY_RE = re.compile(r"^\s*([a-zA-Z0-9_]+:[^\s]+)\s+(.+?)\s*$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        default="GSE172268_merge_Annotation_Result.xlsx",
        help="Path to annotation workbook.",
    )
    parser.add_argument(
        "--sheet",
        default="expression_profile",
        help="Worksheet name with expression and annotation columns.",
    )
    parser.add_argument(
        "--output-dir",
        default="netact_input",
        help="Directory for NetAct-ready output files.",
    )
    return parser.parse_args()


def normalize_symbol(raw: str | None) -> str | None:
    if not raw:
        return None
    value = raw.strip().strip(",;|")
    if not value or value == ".":
        return None
    if value.lower().startswith("loc") and value[3:].isdigit():
        return value.upper()
    cleaned = re.sub(r"[^0-9A-Za-z._-]", "", value)
    return cleaned if cleaned else None


def parse_ko_annotation(raw: str | None) -> tuple[str | None, str]:
    if not raw:
        return None, ""
    text = str(raw).strip()
    if not text or text == ".":
        return None, ""
    match = KO_ENTRY_RE.match(text)
    if match:
        remainder = match.group(2).strip()
        if ";" in remainder:
            symbol_part, _ = remainder.split(";", 1)
            symbol_part = symbol_part.strip()
            if " " not in symbol_part:
                return normalize_symbol(symbol_part), text.lower()
    return None, text.lower()


def is_tf(symbol: str | None, annotation_text: str) -> bool:
    if symbol and TF_SYMBOL_RE.search(symbol):
        return True
    return any(keyword in annotation_text for keyword in TF_KEYWORDS)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found in {input_path.name}")
    ws = wb[args.sheet]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = [str(c).strip() if c is not None else "" for c in header]
    col_index = {name: idx for idx, name in enumerate(columns)}

    contig_idx = col_index.get("Contig")
    if contig_idx is None:
        raise ValueError("Required column 'Contig' not found.")

    expr_columns = [name for name in columns if name.endswith("_FPKM")]
    if not expr_columns:
        raise ValueError("No expression columns ending with '_FPKM' were found.")

    annotation_col = columns[-1]
    annotation_idx = len(columns) - 1

    sums_by_gene: dict[str, list[float]] = {}
    counts_by_gene: dict[str, int] = defaultdict(int)
    contigs_by_gene: dict[str, set[str]] = defaultdict(set)
    tf_by_gene: dict[str, bool] = defaultdict(bool)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        contig = str(row[contig_idx]).strip() if row[contig_idx] else ""
        symbol, annotation_text = parse_ko_annotation(row[annotation_idx])
        gene_id = symbol or contig
        if not gene_id:
            continue

        values: list[float] = []
        for name in expr_columns:
            idx = col_index[name]
            raw = row[idx] if idx < len(row) else 0
            values.append(float(raw) if raw is not None else 0.0)

        if gene_id not in sums_by_gene:
            sums_by_gene[gene_id] = [0.0] * len(expr_columns)
        for i, value in enumerate(values):
            sums_by_gene[gene_id][i] += value
        counts_by_gene[gene_id] += 1
        if contig:
            contigs_by_gene[gene_id].add(contig)
        tf_by_gene[gene_id] = tf_by_gene[gene_id] or is_tf(symbol, annotation_text)

    expression_path = output_dir / "expression_matrix.tsv"
    with expression_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["gene"] + expr_columns)
        for gene in sorted(sums_by_gene):
            count = counts_by_gene[gene]
            mean_values = [f"{v / count:.6f}" for v in sums_by_gene[gene]]
            writer.writerow([gene] + mean_values)

    tf_path = output_dir / "tf_list.txt"
    with tf_path.open("w", encoding="utf-8") as f:
        for gene in sorted(g for g, is_tf_gene in tf_by_gene.items() if is_tf_gene):
            f.write(f"{gene}\n")

    metadata_path = output_dir / "gene_metadata.tsv"
    with metadata_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["gene", "contig_count", "contigs", "is_tf", "annotation_column"])
        for gene in sorted(sums_by_gene):
            contigs = sorted(contigs_by_gene.get(gene, set()))
            writer.writerow(
                [
                    gene,
                    len(contigs),
                    ",".join(contigs),
                    "1" if tf_by_gene.get(gene, False) else "0",
                    annotation_col,
                ]
            )

    print(f"Wrote {expression_path}")
    print(f"Wrote {tf_path}")
    print(f"Wrote {metadata_path}")
    print(f"Rows processed into {len(sums_by_gene)} unique genes.")


if __name__ == "__main__":
    main()
